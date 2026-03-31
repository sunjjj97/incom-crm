from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Q, Count
from django.utils import timezone
from django.utils.safestring import mark_safe
from django.conf import settings
import pytz
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from .models import Department, Team, TeamMember, ContractCompany, LoginLog
from django.contrib.auth.models import User
import pandas as pd
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from functools import wraps
from django.core.exceptions import PermissionDenied


def staff_required(view_func):
    """관리자 권한 필요 데코레이터 (읽기 전용 사용자는 차단)"""
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        # is_staff 또는 is_superuser만 허용 (읽기 전용 사용자는 is_staff=False, is_superuser=False)
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, '관리자 권한이 필요합니다.')
            raise PermissionDenied
        return view_func(request, *args, **kwargs)
    return _wrapped_view


def login_view(request):
    """로그인 뷰"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    # IP 주소 및 User-Agent 가져오기
    ip_address = request.META.get('REMOTE_ADDR')
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # 활성화된 사용자인지 확인 (is_active=True)
            if user.is_active:
                login(request, user)
                # 로그인 성공 로그 기록
                LoginLog.objects.create(
                    user=user,
                    username=username,
                    ip_address=ip_address,
                    user_agent=user_agent,
                    success=True
                )
                return redirect('dashboard')
            else:
                # 비활성화된 계정 로그 기록
                LoginLog.objects.create(
                    username=username,
                    ip_address=ip_address,
                    user_agent=user_agent,
                    success=False,
                    failure_reason='비활성화된 계정'
                )
                messages.error(request, '비활성화된 계정입니다.')
        else:
            # 로그인 실패 로그 기록
            LoginLog.objects.create(
                username=username,
                ip_address=ip_address,
                user_agent=user_agent,
                success=False,
                failure_reason='아이디 또는 비밀번호 오류'
            )
            messages.error(request, '아이디 또는 비밀번호가 올바르지 않습니다.')
    
    return render(request, 'crm_app/login.html')


def logout_view(request):
    """로그아웃 뷰"""
    logout(request)
    return redirect('login')


@login_required
def dashboard(request):
    """대시보드 - 월별 매출 그래프"""
    # 필터 파라미터 처리
    filter_department_id = request.GET.get('department', None)
    filter_team_id = request.GET.get('team', None)
    filter_date = request.GET.get('date', None)  # YYYY-MM 형식
    
    # 날짜 파싱
    if filter_date:
        try:
            filter_year, filter_month = map(int, filter_date.split('-'))
        except:
            filter_year = None
            filter_month = None
    else:
        filter_year = None
        filter_month = None
    
    # 현재 년도와 월 (필터가 없으면 현재 날짜 사용)
    current_year = filter_year if filter_year else timezone.now().year
    current_month = filter_month if filter_month else timezone.now().month
    
    # 기본 쿼리셋 (필터 적용)
    base_query = ContractCompany.objects.all()
    
    if filter_department_id:
        base_query = base_query.filter(team__department_id=filter_department_id)
    
    if filter_team_id:
        base_query = base_query.filter(team_id=filter_team_id)
    
    team_members = TeamMember.objects.all()
    
    # 당월 개인실적 계산 (단위: 만원)
    current_month_personal_performance = []
    for member in team_members:
        query = base_query.filter(
            team_member=member,
            contract_date__year=current_year,
            contract_date__month=current_month
        )
        result = query.aggregate(total=Sum('payment_amount'))
        total = float(result['total']) if result['total'] is not None else 0.0
        # 만원 단위로 변환
        total_in_manwon = total / 10000
        if total_in_manwon > 0:
            current_month_personal_performance.append({
                'name': member.name,
                'value': total_in_manwon
            })
    # 높은 순서대로 정렬
    current_month_personal_performance.sort(key=lambda x: x['value'], reverse=True)
    
    # 누적 개인실적 계산 (단위: 만원, 현재 년도 기준)
    cumulative_personal_performance = []
    for member in team_members:
        query = base_query.filter(
            team_member=member,
            contract_date__year=current_year
        )
        result = query.aggregate(total=Sum('payment_amount'))
        total = float(result['total']) if result['total'] is not None else 0.0
        # 만원 단위로 변환
        total_in_manwon = total / 10000
        if total_in_manwon > 0:
            cumulative_personal_performance.append({
                'name': member.name,
                'value': total_in_manwon
            })
    # 높은 순서대로 정렬
    cumulative_personal_performance.sort(key=lambda x: x['value'], reverse=True)
    
    # 필터별 월별 실적 계산 (부서/팀/개인) - 건수 포함
    # 부서실적 (현재 월 기준)
    department_performance = []
    departments_for_perf = Department.objects.all()
    if filter_department_id:
        departments_for_perf = departments_for_perf.filter(id=filter_department_id)
    for dept in departments_for_perf:
        dept_query = base_query.filter(
            team__department=dept,
            contract_date__year=current_year,
            contract_date__month=current_month
        )
        result = dept_query.aggregate(total=Sum('payment_amount'), count=Count('id'))
        total = float(result['total']) if result['total'] is not None else 0.0
        count = result['count'] if result['count'] is not None else 0
        if total > 0:
            department_performance.append({
                'name': dept.name,
                'value': total,
                'count': count
            })
    department_performance.sort(key=lambda x: x['value'], reverse=True)
    
    # 팀실적 (현재 월 기준)
    team_performance = []
    teams_for_perf = Team.objects.all()
    if filter_team_id:
        teams_for_perf = teams_for_perf.filter(id=filter_team_id)
    elif filter_department_id:
        teams_for_perf = teams_for_perf.filter(department_id=filter_department_id)
    for team in teams_for_perf:
        team_query = base_query.filter(
            team=team,
            contract_date__year=current_year,
            contract_date__month=current_month
        )
        result = team_query.aggregate(total=Sum('payment_amount'), count=Count('id'))
        total = float(result['total']) if result['total'] is not None else 0.0
        count = result['count'] if result['count'] is not None else 0
        if total > 0:
            team_performance.append({
                'name': team.name,
                'value': total,
                'count': count
            })
    team_performance.sort(key=lambda x: x['value'], reverse=True)
    
    # 개인실적 (현재 월 기준) - 건수 포함
    personal_performance_text = []
    for member in team_members:
        query = base_query.filter(
            team_member=member,
            contract_date__year=current_year,
            contract_date__month=current_month
        )
        result = query.aggregate(total=Sum('payment_amount'), count=Count('id'))
        total = float(result['total']) if result['total'] is not None else 0.0
        count = result['count'] if result['count'] is not None else 0
        if total > 0:
            personal_performance_text.append({
                'name': member.name,
                'value': total,
                'count': count
            })
    personal_performance_text.sort(key=lambda x: x['value'], reverse=True)
    
    # 실적 타입 선택 (기본값: 팀실적)
    performance_type = request.GET.get('performance_type', 'team')  # 'department', 'team', 'personal'
    
    # 합계 계산
    dept_total_amount = sum([d['value'] for d in department_performance])
    dept_total_count = sum([d['count'] for d in department_performance])
    team_total_amount = sum([t['value'] for t in team_performance])
    team_total_count = sum([t['count'] for t in team_performance])
    personal_total_amount = sum([p['value'] for p in personal_performance_text])
    personal_total_count = sum([p['count'] for p in personal_performance_text])
    
    # 통계 데이터
    total_result = ContractCompany.objects.aggregate(total=Sum('payment_amount'))
    total_sales = float(total_result['total']) if total_result['total'] is not None else 0.0
    total_companies = ContractCompany.objects.count()
    monthly_result = ContractCompany.objects.filter(
        contract_date__year=current_year,
        contract_date__month=current_month
    ).aggregate(total=Sum('payment_amount'))
    monthly_sales = float(monthly_result['total']) if monthly_result['total'] is not None else 0.0
    
    # 당일 매출 계산 (한국 시간 기준 0:00를 기준으로 날짜 변경)
    # 한국 시간대(Asia/Seoul)로 변환하여 오늘 날짜 계산
    korea_tz = pytz.timezone('Asia/Seoul')
    korea_now = timezone.now().astimezone(korea_tz)
    today = korea_now.date()  # 한국 시간 기준 오늘 날짜
    
    daily_result = ContractCompany.objects.filter(
        contract_date=today
    ).aggregate(total=Sum('payment_amount'))
    daily_sales = float(daily_result['total']) if daily_result['total'] is not None else 0.0
    
    # 당일 계약건 목록 (업체명, 담당자, 금액)
    today_contracts = ContractCompany.objects.filter(
        contract_date=today
    ).select_related('team_member', 'team').order_by('-created_at')
    
    # 필터용 데이터
    departments = Department.objects.all().order_by('name')
    teams_filter = Team.objects.all().order_by('name')
    
    context = {
        'current_month_personal_performance': current_month_personal_performance,
        'cumulative_personal_performance': cumulative_personal_performance,
        'total_sales': total_sales,
        'total_companies': total_companies,
        'monthly_sales': monthly_sales,
        'daily_sales': daily_sales,
        'today_contracts': today_contracts,
        'today': today,
        'current_year': current_year,
        'current_month': current_month,
        'departments': departments,
        'teams_filter': teams_filter,
        'filter_department_id': int(filter_department_id) if filter_department_id else None,
        'filter_team_id': int(filter_team_id) if filter_team_id else None,
        'filter_date': filter_date or f"{current_year}-{current_month:02d}",
        'department_performance': department_performance,
        'team_performance': team_performance,
        'personal_performance_text': personal_performance_text,
        'performance_type': performance_type,
        'dept_total_amount': dept_total_amount,
        'dept_total_count': dept_total_count,
        'team_total_amount': team_total_amount,
        'team_total_count': team_total_count,
        'personal_total_amount': personal_total_amount,
        'personal_total_count': personal_total_count,
    }
    
    return render(request, 'crm_app/dashboard.html', context)


@login_required
def sales_list(request):
    """전체 매출 리스트"""
    # 데이터베이스에 contract_type 컬럼이 있는지 확인
    from django.db import connection
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name='crm_app_contractcompany' AND column_name='contract_type'
            """)
            has_contract_type = cursor.fetchone() is not None
    except Exception:
        # PostgreSQL이 아닌 경우 또는 에러 발생 시 기본값 사용
        has_contract_type = False
    
    if has_contract_type:
        sales = ContractCompany.objects.all().select_related('team_member', 'team').order_by('-created_at')
    else:
        # contract_type 컬럼이 없으면 기본 필드만 사용 (마이그레이션 전)
        sales = ContractCompany.objects.only(
            'id', 'company_name', 'representative', 'phone', 'contract_date', 
            'contract_expiry', 'payment_amount', 'team_member_id', 'team_id', 
            'notes', 'created_at', 'updated_at'
        ).select_related('team_member', 'team').order_by('-created_at')
    
    # 필터링
    search = request.GET.get('search', '')
    team_filter = request.GET.get('team', '')
    member_filter = request.GET.get('member', '')
    year_filter = request.GET.get('year', '')
    month_filter = request.GET.get('month', '')
    show_duplicates = request.GET.get('duplicates', '')
    
    if search:
        sales = sales.filter(
            Q(company_name__icontains=search) |
            Q(representative__icontains=search) |
            Q(phone__icontains=search)
        )
    
    if team_filter:
        sales = sales.filter(team_id=team_filter)
    
    if member_filter:
        sales = sales.filter(team_member_id=member_filter)
    
    if year_filter:
        sales = sales.filter(contract_date__year=year_filter)
    
    if month_filter:
        sales = sales.filter(contract_date__month=month_filter)
    
    # 중복 업체 필터링
    if show_duplicates == 'true':
        # 업체명별 그룹화하여 2개 이상인 것만 필터링
        from django.db.models import Count
        duplicate_companies = sales.values('company_name').annotate(
            count=Count('id')
        ).filter(count__gt=1).values_list('company_name', flat=True)
        sales = sales.filter(company_name__in=duplicate_companies)
    
    # 총 매출 계산
    total_result = sales.aggregate(total=Sum('payment_amount'))
    total_sales = float(total_result['total']) if total_result['total'] is not None else 0.0
    
    # 중복 업체 통계 계산
    from django.db.models import Count
    duplicate_stats = ContractCompany.objects.values('company_name').annotate(
        count=Count('id')
    ).filter(count__gt=1).order_by('-count')
    duplicate_count = duplicate_stats.count()
    
    # 페이지네이션
    from django.core.paginator import Paginator
    paginator = Paginator(sales, 20)  # 페이지당 20개
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'sales': page_obj,
        'page_obj': page_obj,
        'teams': Team.objects.all(),
        'members': TeamMember.objects.all(),
        'total_sales': total_sales,
        'search': search,
        'team_filter': team_filter,
        'member_filter': member_filter,
        'year_filter': year_filter,
        'month_filter': month_filter,
        'show_duplicates': show_duplicates,
        'duplicate_count': duplicate_count,
        'duplicate_stats': duplicate_stats[:10],  # 상위 10개만 표시
    }
    
    return render(request, 'crm_app/sales_list.html', context)


@staff_required
@login_required
def sales_add(request):
    """매출 등록"""
    if request.method == 'POST':
        company_name = request.POST.get('company_name')
        representative = request.POST.get('representative') or None  # 빈 문자열을 None으로 변환
        phone = request.POST.get('phone') or None  # 빈 문자열을 None으로 변환
        contract_date = request.POST.get('contract_date')
        contract_expiry = request.POST.get('contract_expiry')
        payment_amount = request.POST.get('payment_amount')
        team_member_id = request.POST.get('team_member')
        team_id = request.POST.get('team')
        notes = request.POST.get('notes', '')
        
        team_member = None
        if team_member_id:
            team_member = get_object_or_404(TeamMember, id=team_member_id)
        
        team = None
        if team_id:
            team = get_object_or_404(Team, id=team_id)
        
        # payment_amount를 Decimal로 변환 (쉼표 제거)
        if payment_amount:
            payment_amount = str(payment_amount).replace(',', '').strip()
            try:
                payment_amount = Decimal(payment_amount)
            except (ValueError, InvalidOperation):
                payment_amount = Decimal('0')
        else:
            payment_amount = Decimal('0')
        
        # 계약 유형 및 이전 계약 처리
        contract_type = request.POST.get('contract_type', 'new')
        previous_contract_id = request.POST.get('previous_contract', '')
        previous_contract = None
        
        if previous_contract_id:
            try:
                previous_contract = ContractCompany.objects.get(id=previous_contract_id)
            except ContractCompany.DoesNotExist:
                pass
        
        ContractCompany.objects.create(
            company_name=company_name,
            representative=representative,
            phone=phone,
            contract_date=contract_date,
            contract_expiry=contract_expiry,
            payment_amount=payment_amount,
            team_member=team_member,
            team=team,
            notes=notes,
            contract_type=contract_type,
            previous_contract=previous_contract
        )
        
        messages.success(request, '매출이 등록되었습니다.')
        return redirect('sales_list')
    
    # 같은 업체명의 이전 계약 목록 (계약 유형 선택용)
    companies = ContractCompany.objects.all().order_by('-created_at')
    
    context = {
        'teams': Team.objects.all(),
        'members': TeamMember.objects.all(),
        'companies': companies,
    }
    
    return render(request, 'crm_app/sales_add.html', context)


@staff_required
@login_required
def sales_edit(request, pk):
    """매출 수정"""
    sale = get_object_or_404(ContractCompany, pk=pk)
    
    if request.method == 'POST':
        sale.company_name = request.POST.get('company_name')
        sale.representative = request.POST.get('representative') or None  # 빈 문자열을 None으로 변환
        sale.phone = request.POST.get('phone') or None  # 빈 문자열을 None으로 변환
        sale.contract_date = request.POST.get('contract_date')
        sale.contract_expiry = request.POST.get('contract_expiry')
        
        # payment_amount를 Decimal로 변환 (쉼표 제거)
        payment_amount = request.POST.get('payment_amount')
        if payment_amount:
            payment_amount = str(payment_amount).replace(',', '').strip()
            try:
                sale.payment_amount = Decimal(payment_amount)
            except (ValueError, InvalidOperation):
                sale.payment_amount = Decimal('0')
        else:
            sale.payment_amount = Decimal('0')
        
        team_member_id = request.POST.get('team_member')
        if team_member_id:
            sale.team_member = get_object_or_404(TeamMember, id=team_member_id)
        else:
            sale.team_member = None
        
        team_id = request.POST.get('team')
        if team_id:
            sale.team = get_object_or_404(Team, id=team_id)
        else:
            sale.team = None
        
        sale.notes = request.POST.get('notes', '')
        
        # 계약 유형 및 이전 계약 처리
        contract_type = request.POST.get('contract_type', 'new')
        previous_contract_id = request.POST.get('previous_contract', '')
        
        sale.contract_type = contract_type
        if previous_contract_id:
            try:
                sale.previous_contract = ContractCompany.objects.get(id=previous_contract_id)
            except ContractCompany.DoesNotExist:
                sale.previous_contract = None
        else:
            sale.previous_contract = None
        
        sale.save()
        
        messages.success(request, '매출이 수정되었습니다.')
        return redirect('sales_list')
    
    # 같은 업체명의 다른 계약 목록 (계약 유형 선택용)
    companies = ContractCompany.objects.exclude(id=sale.id).order_by('-created_at')
    
    context = {
        'sale': sale,
        'teams': Team.objects.all(),
        'members': TeamMember.objects.all(),
        'companies': companies,
    }
    
    return render(request, 'crm_app/sales_edit.html', context)


@staff_required
@login_required
def sales_delete(request, pk):
    """매출 삭제"""
    sale = get_object_or_404(ContractCompany, pk=pk)
    
    if request.method == 'POST':
        sale.delete()
        messages.success(request, '매출이 삭제되었습니다.')
        return redirect('sales_list')
    
    return render(request, 'crm_app/sales_delete.html', {'sale': sale})


@login_required
def company_list(request):
    """계약업체 리스트"""
    # 데이터베이스에 contract_type 컬럼이 있는지 확인
    from django.db import connection
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name='crm_app_contractcompany' AND column_name='contract_type'
            """)
            has_contract_type = cursor.fetchone() is not None
    except Exception:
        has_contract_type = False
    
    if has_contract_type:
        companies = ContractCompany.objects.all().select_related('team_member', 'team').order_by('-created_at')
    else:
        # contract_type 컬럼이 없으면 기본 필드만 사용
        companies = ContractCompany.objects.only(
            'id', 'company_name', 'representative', 'phone', 'contract_date', 
            'contract_expiry', 'payment_amount', 'team_member_id', 'team_id', 
            'notes', 'created_at', 'updated_at'
        ).select_related('team_member', 'team').order_by('-created_at')
    
    # 필터링
    search = request.GET.get('search', '')
    team_filter = request.GET.get('team', '')
    show_duplicates = request.GET.get('duplicates', '')
    
    if search:
        companies = companies.filter(
            Q(company_name__icontains=search) |
            Q(representative__icontains=search) |
            Q(phone__icontains=search)
        )
    
    if team_filter:
        companies = companies.filter(team_id=team_filter)
    
    # 중복 업체 필터링
    if show_duplicates == 'true':
        # 업체명별 그룹화하여 2개 이상인 것만 필터링
        from django.db.models import Count
        duplicate_companies = companies.values('company_name').annotate(
            count=Count('id')
        ).filter(count__gt=1).values_list('company_name', flat=True)
        companies = companies.filter(company_name__in=duplicate_companies)
    
    # 중복 업체 통계 계산
    from django.db.models import Count
    duplicate_stats = ContractCompany.objects.values('company_name').annotate(
        count=Count('id')
    ).filter(count__gt=1).order_by('-count')
    duplicate_count = duplicate_stats.count()
    
    # 페이지네이션
    from django.core.paginator import Paginator
    paginator = Paginator(companies, 20)  # 페이지당 20개
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'companies': page_obj,
        'page_obj': page_obj,
        'teams': Team.objects.all(),
        'search': search,
        'team_filter': team_filter,
        'show_duplicates': show_duplicates,
        'duplicate_count': duplicate_count,
        'duplicate_stats': duplicate_stats[:10],  # 상위 10개만 표시
    }
    
    return render(request, 'crm_app/company_list.html', context)


@staff_required
@login_required
def duplicate_companies(request):
    """중복 업체 관리 페이지"""
    company_name = request.GET.get('name', '')
    
    if not company_name:
        # 중복 업체 목록 표시
        duplicate_stats = ContractCompany.objects.values('company_name').annotate(
            count=Count('id')
        ).filter(count__gt=1).order_by('-count')
        
        context = {
            'duplicate_stats': duplicate_stats,
        }
        return render(request, 'crm_app/duplicate_companies_list.html', context)
    
    # 특정 업체명의 중복 항목들 조회 (등록 순서대로)
    duplicates = ContractCompany.objects.filter(
        company_name=company_name
    ).select_related('team_member', 'team', 'previous_contract').order_by('-created_at')
    
    # 계약 유형 자동 판단 함수
    def detect_contract_type(contract):
        """계약 유형 자동 판단"""
        if contract.previous_contract:
            return contract.contract_type, contract.previous_contract
        
        # 같은 업체명의 이전 계약 찾기
        previous = ContractCompany.objects.filter(
            company_name=company_name,
            created_at__lt=contract.created_at
        ).order_by('-created_at').first()
        
        if previous:
            # 계약일이 이전 계약 만료일 이후면 재계약
            if contract.contract_date > previous.contract_expiry:
                return 'renewal', previous
            # 계약일이 이전 계약 기간 중이면 추가 계약
            elif previous.contract_date <= contract.contract_date <= previous.contract_expiry:
                return 'additional', previous
        
        return 'new', None
    
    # 각 계약의 추천 유형 계산 (아직 설정되지 않은 경우)
    for duplicate in duplicates:
        if not duplicate.contract_type or duplicate.contract_type == 'new' or not duplicate.previous_contract:
            recommended_type, recommended_previous = detect_contract_type(duplicate)
            # 동적 속성으로 추가
            duplicate.recommended_type = recommended_type
            duplicate.recommended_previous = recommended_previous
    
    context = {
        'company_name': company_name,
        'duplicates': duplicates,
        'count': duplicates.count(),
    }
    
    return render(request, 'crm_app/duplicate_companies_detail.html', context)


@staff_required
@login_required
def duplicate_companies_delete(request):
    """중복 업체 일괄 삭제"""
    if request.method == 'POST':
        company_name = request.POST.get('company_name', '')
        keep_id = request.POST.get('keep_id', '')  # 유지할 항목 ID
        
        if not company_name or not keep_id:
            messages.error(request, '유지할 항목을 선택해주세요.')
            return redirect('duplicate_companies')
        
        try:
            keep_id = int(keep_id)
            # 같은 업체명의 모든 항목 중 keep_id를 제외하고 삭제
            deleted_count = ContractCompany.objects.filter(
                company_name=company_name
            ).exclude(id=keep_id).delete()[0]
            
            messages.success(request, f'{deleted_count}개의 중복 항목이 삭제되었습니다.')
            return redirect('duplicate_companies')
        except Exception as e:
            messages.error(request, f'삭제 중 오류가 발생했습니다: {str(e)}')
            return redirect('duplicate_companies')
    
    return redirect('duplicate_companies')


@login_required
def get_team_members(request):
    """팀 선택 시 해당 팀의 팀원 목록을 반환하는 API"""
    team_id = request.GET.get('team_id')
    
    if not team_id:
        return JsonResponse({'members': []})
    
    try:
        members = TeamMember.objects.filter(team_id=team_id).order_by('name')
        members_data = [{'id': member.id, 'name': member.name} for member in members]
        return JsonResponse({'members': members_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@staff_required
@login_required
def duplicate_companies_link(request):
    """중복 업체 계약 연결 (재계약/추가계약 설정)"""
    if request.method == 'POST':
        contract_id = request.POST.get('contract_id')
        contract_type = request.POST.get('contract_type')
        previous_contract_id = request.POST.get('previous_contract_id', '')
        
        if not contract_id or not contract_type:
            messages.error(request, '필수 정보가 누락되었습니다.')
            return redirect('duplicate_companies')
        
        try:
            contract = ContractCompany.objects.get(id=contract_id)
            contract.contract_type = contract_type
            
            if previous_contract_id:
                previous_contract = ContractCompany.objects.get(id=previous_contract_id)
                contract.previous_contract = previous_contract
            
            contract.save()
            messages.success(request, f'계약 유형이 "{contract.get_contract_type_display()}"로 설정되었습니다.')
            
            # 원래 페이지로 리다이렉트
            company_name = request.POST.get('company_name', '')
            if company_name:
                return redirect(f'{reverse("duplicate_companies")}?name={company_name}')
            return redirect('duplicate_companies')
        except ContractCompany.DoesNotExist:
            messages.error(request, '계약을 찾을 수 없습니다.')
            return redirect('duplicate_companies')
        except Exception as e:
            messages.error(request, f'처리 중 오류가 발생했습니다: {str(e)}')
            return redirect('duplicate_companies')
    
    return redirect('duplicate_companies')


@login_required
def team_member_list(request):
    """팀원 리스트"""
    members = TeamMember.objects.all().select_related('team')
    
    # 필터링
    team_filter = request.GET.get('team', '')
    if team_filter:
        members = members.filter(team_id=team_filter)
    
    context = {
        'members': members,
        'teams': Team.objects.all(),
        'team_filter': team_filter,
    }
    
    return render(request, 'crm_app/team_member_list.html', context)


@staff_required
@login_required
def team_member_add(request):
    """팀원 등록"""
    if request.method == 'POST':
        name = request.POST.get('name')
        team_id = request.POST.get('team')
        email = request.POST.get('email', '')
        phone = request.POST.get('phone', '')
        
        team = None
        if team_id:
            team = get_object_or_404(Team, id=team_id)
        
        TeamMember.objects.create(
            name=name,
            team=team,
            email=email,
            phone=phone
        )
        
        messages.success(request, '팀원이 등록되었습니다.')
        return redirect('team_member_list')
    
    context = {
        'teams': Team.objects.all(),
    }
    
    return render(request, 'crm_app/team_member_add.html', context)


@staff_required
@login_required
def team_member_edit(request, pk):
    """팀원 수정"""
    member = get_object_or_404(TeamMember, pk=pk)
    
    if request.method == 'POST':
        member.name = request.POST.get('name')
        member.email = request.POST.get('email', '')
        member.phone = request.POST.get('phone', '')
        
        team_id = request.POST.get('team')
        if team_id:
            member.team = get_object_or_404(Team, id=team_id)
        else:
            member.team = None
        
        member.save()
        
        messages.success(request, '팀원 정보가 수정되었습니다.')
        return redirect('team_member_list')
    
    context = {
        'member': member,
        'teams': Team.objects.all(),
    }
    
    return render(request, 'crm_app/team_member_edit.html', context)


@staff_required
@login_required
def team_member_delete(request, pk):
    """팀원 삭제"""
    member = get_object_or_404(TeamMember, pk=pk)
    
    if request.method == 'POST':
        member.delete()
        messages.success(request, '팀원이 삭제되었습니다.')
        return redirect('team_member_list')
    
    return render(request, 'crm_app/team_member_delete.html', {'member': member})


# 팀 관리
@login_required
def team_list(request):
    """팀 리스트"""
    teams = Team.objects.all().select_related('department')
    
    # 필터링
    dept_filter = request.GET.get('department', '')
    if dept_filter:
        teams = teams.filter(department_id=dept_filter)
    
    context = {
        'teams': teams,
        'departments': Department.objects.all(),
        'dept_filter': dept_filter,
    }
    
    return render(request, 'crm_app/team_list.html', context)


@staff_required
@login_required
def team_add(request):
    """팀 등록"""
    if request.method == 'POST':
        name = request.POST.get('name')
        department_id = request.POST.get('department')
        
        department = None
        if department_id:
            department = get_object_or_404(Department, id=department_id)
        
        Team.objects.create(
            name=name,
            department=department
        )
        
        messages.success(request, '팀이 등록되었습니다.')
        return redirect('team_list')
    
    context = {
        'departments': Department.objects.all(),
    }
    
    return render(request, 'crm_app/team_add.html', context)


@staff_required
@login_required
def team_edit(request, pk):
    """팀 수정"""
    team = get_object_or_404(Team, pk=pk)
    
    if request.method == 'POST':
        team.name = request.POST.get('name')
        
        department_id = request.POST.get('department')
        if department_id:
            team.department = get_object_or_404(Department, id=department_id)
        else:
            team.department = None
        
        team.save()
        
        messages.success(request, '팀 정보가 수정되었습니다.')
        return redirect('team_list')
    
    context = {
        'team': team,
        'departments': Department.objects.all(),
    }
    
    return render(request, 'crm_app/team_edit.html', context)


@staff_required
@login_required
def team_delete(request, pk):
    """팀 삭제"""
    team = get_object_or_404(Team, pk=pk)
    
    if request.method == 'POST':
        team.delete()
        messages.success(request, '팀이 삭제되었습니다.')
        return redirect('team_list')
    
    return render(request, 'crm_app/team_delete.html', {'team': team})


# 사업부 관리
@login_required
def department_list(request):
    """사업부 리스트"""
    departments = Department.objects.all()
    
    context = {
        'departments': departments,
    }
    
    return render(request, 'crm_app/department_list.html', context)


@staff_required
@login_required
def department_add(request):
    """사업부 등록"""
    if request.method == 'POST':
        name = request.POST.get('name')
        
        Department.objects.create(name=name)
        
        messages.success(request, '사업부가 등록되었습니다.')
        return redirect('department_list')
    
    return render(request, 'crm_app/department_add.html')


@staff_required
@login_required
def department_edit(request, pk):
    """사업부 수정"""
    department = get_object_or_404(Department, pk=pk)
    
    if request.method == 'POST':
        department.name = request.POST.get('name')
        department.save()
        
        messages.success(request, '사업부 정보가 수정되었습니다.')
        return redirect('department_list')
    
    return render(request, 'crm_app/department_edit.html', {'department': department})


@staff_required
@login_required
def department_delete(request, pk):
    """사업부 삭제"""
    department = get_object_or_404(Department, pk=pk)
    
    if request.method == 'POST':
        department.delete()
        messages.success(request, '사업부가 삭제되었습니다.')
        return redirect('department_list')
    
    return render(request, 'crm_app/department_delete.html', {'department': department})


@staff_required
@login_required
def export_excel(request):
    """엑셀 백업"""
    # 계약업체 데이터
    companies = ContractCompany.objects.all().select_related('team_member', 'team', 'team__department')
    
    data = []
    for company in companies:
        data.append({
            '업체명': company.company_name,
            '대표자명': company.representative,
            '전화번호': company.phone,
            '계약일': company.contract_date.strftime('%Y-%m-%d'),
            '계약만료일': company.contract_expiry.strftime('%Y-%m-%d'),
            '결제금액': float(company.payment_amount),
            '담당 팀원': company.team_member.name if company.team_member else '',
            '담당 팀': company.team.name if company.team else '',
            '사업부': company.team.department.name if company.team and company.team.department else '',
            '비고': company.notes or '',
            '생성일': company.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })
    
    df = pd.DataFrame(data)
    
    # 팀원 데이터
    members = TeamMember.objects.all().select_related('team', 'team__department')
    member_data = []
    for member in members:
        member_data.append({
            '이름': member.name,
            '팀': member.team.name if member.team else '',
            '사업부': member.team.department.name if member.team and member.team.department else '',
            '이메일': member.email or '',
            '전화번호': member.phone or '',
            '생성일': member.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })
    
    df_members = pd.DataFrame(member_data)
    
    # 엑셀 파일 생성
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='계약업체', index=False)
        df_members.to_excel(writer, sheet_name='팀원', index=False)
    
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="crm_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


@staff_required
@login_required
def sales_template_download(request):
    """엑셀 업로드 템플릿 다운로드"""
    # 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "매출 일괄 등록"
    
    # 헤더 설정
    headers = ['업체명', '대표자명', '전화번호', '계약일', '계약만료일', '결제금액', '담당팀원', '담당팀', '비고']
    ws.append(headers)
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 예시 데이터
    example_data = [
        ['회사A', '홍길동', '010-1234-5678', '2024-01-15', '2024-12-31', '1000000', '신현준', '1팀', '예시 데이터'],
        ['회사B', '김철수', '010-2345-6789', '2024-02-20', '2024-12-31', '2000000', '이현우', '2팀', ''],
    ]
    for row in example_data:
        ws.append(row)
    
    # 컬럼 너비 조정
    column_widths = [20, 15, 18, 12, 12, 15, 15, 15, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 메모 추가
    ws['A11'] = '※ 필수 입력 항목: 업체명, 계약일, 계약만료일, 결제금액'
    ws['A12'] = '※ 담당팀원, 담당팀은 이름으로 입력 (팀원 목록에 등록된 이름)'
    ws['A13'] = '※ 계약일, 계약만료일은 YYYY-MM-DD 형식으로 입력 (예: 2024-01-15)'
    ws['A14'] = '※ 결제금액은 숫자만 입력 (예: 1000000)'
    
    # 응답 생성
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="매출_일괄등록_템플릿.xlsx"'
    
    return response


@staff_required
@login_required
def sales_bulk_upload(request):
    """엑셀 일괄 등록"""
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, '엑셀 파일을 선택해주세요.')
            return render(request, 'crm_app/sales_bulk_upload.html', {
                'teams': Team.objects.all(),
                'members': TeamMember.objects.all(),
            })
        
        excel_file = request.FILES['excel_file']
        
        try:
            # 엑셀 파일 읽기
            df = pd.read_excel(excel_file, engine='openpyxl')
            
            # 필요한 컬럼 확인
            required_columns = ['업체명', '계약일', '계약만료일', '결제금액']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                messages.error(request, f'필수 컬럼이 없습니다: {", ".join(missing_columns)}')
                return render(request, 'crm_app/sales_bulk_upload.html', {
                    'teams': Team.objects.all(),
                    'members': TeamMember.objects.all(),
                })
            
            # 결과 저장
            success_count = 0
            error_count = 0
            errors = []
            
            # 팀원 및 팀 이름 매핑 (캐싱)
            member_name_map = {member.name: member for member in TeamMember.objects.all()}
            team_name_map = {team.name: team for team in Team.objects.all()}
            
            for index, row in df.iterrows():
                try:
                    # 필수 필드 검증
                    company_name = str(row.get('업체명', '')).strip()
                    contract_date_str = str(row.get('계약일', '')).strip()
                    contract_expiry_str = str(row.get('계약만료일', '')).strip()
                    payment_amount = row.get('결제금액', 0)
                    
                    if not company_name:
                        errors.append(f'{index + 2}행: 업체명이 비어있습니다.')
                        error_count += 1
                        continue
                    
                    # 날짜 변환
                    try:
                        if isinstance(contract_date_str, str) and len(contract_date_str) == 10:
                            contract_date = datetime.strptime(contract_date_str, '%Y-%m-%d').date()
                        elif isinstance(contract_date_str, datetime):
                            contract_date = contract_date_str.date()
                        else:
                            contract_date = pd.to_datetime(contract_date_str).date()
                        
                        if isinstance(contract_expiry_str, str) and len(contract_expiry_str) == 10:
                            contract_expiry = datetime.strptime(contract_expiry_str, '%Y-%m-%d').date()
                        elif isinstance(contract_expiry_str, datetime):
                            contract_expiry = contract_expiry_str.date()
                        else:
                            contract_expiry = pd.to_datetime(contract_expiry_str).date()
                    except Exception as e:
                        errors.append(f'{index + 2}행: 날짜 형식 오류 ({contract_date_str}, {contract_expiry_str})')
                        error_count += 1
                        continue
                    
                    # 결제금액 변환
                    try:
                        if pd.isna(payment_amount):
                            payment_amount = Decimal('0')
                        else:
                            payment_amount = Decimal(str(payment_amount).replace(',', ''))
                    except Exception as e:
                        errors.append(f'{index + 2}행: 결제금액 형식 오류 ({payment_amount})')
                        error_count += 1
                        continue
                    
                    # 선택 필드 처리
                    representative = str(row.get('대표자명', '')).strip() or None
                    phone = str(row.get('전화번호', '')).strip() or None
                    notes = str(row.get('비고', '')).strip() or ''
                    
                    # 팀원 및 팀 찾기
                    team_member = None
                    team_member_name = str(row.get('담당팀원', '')).strip()
                    if team_member_name:
                        team_member = member_name_map.get(team_member_name)
                        if not team_member:
                            errors.append(f'{index + 2}행: 담당팀원 "{team_member_name}"을 찾을 수 없습니다. (무시하고 진행)')
                    
                    team = None
                    team_name = str(row.get('담당팀', '')).strip()
                    if team_name:
                        team = team_name_map.get(team_name)
                        if not team:
                            errors.append(f'{index + 2}행: 담당팀 "{team_name}"을 찾을 수 없습니다. (무시하고 진행)')
                    
                    # 데이터 생성
                    ContractCompany.objects.create(
                        company_name=company_name,
                        representative=representative,
                        phone=phone,
                        contract_date=contract_date,
                        contract_expiry=contract_expiry,
                        payment_amount=payment_amount,
                        team_member=team_member,
                        team=team,
                        notes=notes
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f'{index + 2}행: 처리 중 오류 발생 - {str(e)}')
                    error_count += 1
                    continue
            
            # 결과 메시지
            if success_count > 0:
                messages.success(request, f'{success_count}건의 데이터가 성공적으로 등록되었습니다.')
            if error_count > 0:
                error_message = f'{error_count}건의 데이터 처리 중 오류가 발생했습니다.'
                if errors:
                    error_message += '<br>' + '<br>'.join(errors[:10])  # 최대 10개만 표시
                    if len(errors) > 10:
                        error_message += f'<br>... 외 {len(errors) - 10}개 오류'
                messages.warning(request, error_message)
            
            return redirect('sales_list')
            
        except Exception as e:
            messages.error(request, f'엑셀 파일을 읽는 중 오류가 발생했습니다: {str(e)}')
            return render(request, 'crm_app/sales_bulk_upload.html', {
                'teams': Team.objects.all(),
                'members': TeamMember.objects.all(),
            })
    
    return render(request, 'crm_app/sales_bulk_upload.html', {
        'teams': Team.objects.all(),
        'members': TeamMember.objects.all(),
    })


@login_required
def login_log_list(request):
    """로그인 로그 목록"""
    logs = LoginLog.objects.all().select_related('user')
    
    # 필터링
    search = request.GET.get('search', '')
    success_filter = request.GET.get('success', '')
    
    if search:
        logs = logs.filter(
            Q(username__icontains=search) |
            Q(ip_address__icontains=search)
        )
    
    if success_filter:
        logs = logs.filter(success=(success_filter == 'true'))
    
    # 페이지네이션 (선택사항)
    from django.core.paginator import Paginator
    paginator = Paginator(logs, 50)  # 페이지당 50개
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'logs': page_obj,
        'search': search,
        'success_filter': success_filter,
    }
    
    return render(request, 'crm_app/login_log_list.html', context)


@login_required
def change_password(request):
    """사용자 본인 비밀번호 변경"""
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password1 = request.POST.get('new_password1')
        new_password2 = request.POST.get('new_password2')
        
        # 현재 비밀번호 확인
        if not request.user.check_password(old_password):
            messages.error(request, '현재 비밀번호가 올바르지 않습니다.')
            return render(request, 'crm_app/change_password.html')
        
        # 새 비밀번호 확인
        if new_password1 != new_password2:
            messages.error(request, '새 비밀번호가 일치하지 않습니다.')
            return render(request, 'crm_app/change_password.html')
        
        # 비밀번호 유효성 검사
        if len(new_password1) < 4:
            messages.error(request, '비밀번호는 최소 4자 이상이어야 합니다.')
            return render(request, 'crm_app/change_password.html')
        
        # 비밀번호 변경
        request.user.set_password(new_password1)
        request.user.save()
        
        # 비밀번호 변경 후 다시 로그인
        from django.contrib.auth import update_session_auth_hash
        update_session_auth_hash(request, request.user)
        
        messages.success(request, '비밀번호가 성공적으로 변경되었습니다.')
        return redirect('dashboard')
    
    return render(request, 'crm_app/change_password.html')


# 사용자 관리
@staff_required
@login_required
def user_list(request):
    """사용자 리스트"""
    users = User.objects.all().order_by('username')
    
    # 필터링
    search = request.GET.get('search', '')
    is_staff_filter = request.GET.get('is_staff', '')
    is_active_filter = request.GET.get('is_active', '')
    
    if search:
        users = users.filter(
            Q(username__icontains=search) |
            Q(email__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search)
        )
    
    if is_staff_filter:
        users = users.filter(is_staff=(is_staff_filter == 'true'))
    
    if is_active_filter:
        users = users.filter(is_active=(is_active_filter == 'true'))
    
    context = {
        'users': users,
        'search': search,
        'is_staff_filter': is_staff_filter,
        'is_active_filter': is_active_filter,
    }
    
    return render(request, 'crm_app/user_list.html', context)


@staff_required
@login_required
def user_add(request):
    """사용자 등록"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        email = request.POST.get('email', '')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        is_staff = request.POST.get('is_staff') == 'on'
        is_superuser = request.POST.get('is_superuser') == 'on'
        is_active = request.POST.get('is_active') == 'on'
        
        # 유효성 검사
        if User.objects.filter(username=username).exists():
            messages.error(request, '이미 존재하는 아이디입니다.')
            return render(request, 'crm_app/user_add.html')
        
        if password1 != password2:
            messages.error(request, '비밀번호가 일치하지 않습니다.')
            return render(request, 'crm_app/user_add.html')
        
        if len(password1) < 4:
            messages.error(request, '비밀번호는 최소 4자 이상이어야 합니다.')
            return render(request, 'crm_app/user_add.html')
        
        # 사용자 생성
        user = User.objects.create_user(
            username=username,
            password=password1,
            email=email,
            first_name=first_name,
            last_name=last_name,
            is_staff=is_staff,
            is_superuser=is_superuser,
            is_active=is_active
        )
        
        messages.success(request, f'사용자 "{username}"가 생성되었습니다.')
        return redirect('user_list')
    
    return render(request, 'crm_app/user_add.html')


@staff_required
@login_required
def user_edit(request, pk):
    """사용자 수정"""
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email', '')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        is_staff = request.POST.get('is_staff') == 'on'
        is_superuser = request.POST.get('is_superuser') == 'on'
        is_active = request.POST.get('is_active') == 'on'
        
        # 아이디 중복 확인 (본인 제외)
        if User.objects.filter(username=username).exclude(pk=user.pk).exists():
            messages.error(request, '이미 존재하는 아이디입니다.')
            return render(request, 'crm_app/user_edit.html', {'user': user})
        
        user.username = username
        user.email = email
        user.first_name = first_name
        user.last_name = last_name
        user.is_staff = is_staff
        user.is_superuser = is_superuser
        user.is_active = is_active
        user.save()
        
        messages.success(request, '사용자 정보가 수정되었습니다.')
        return redirect('user_list')
    
    return render(request, 'crm_app/user_edit.html', {'user': user})


@staff_required
@login_required
def user_delete(request, pk):
    """사용자 삭제"""
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        username = user.username
        user.delete()
        messages.success(request, f'사용자 "{username}"가 삭제되었습니다.')
        return redirect('user_list')
    
    return render(request, 'crm_app/user_delete.html', {'user': user})


@staff_required
@login_required
def user_change_password(request, pk):
    """관리자가 사용자 비밀번호 변경"""
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        new_password1 = request.POST.get('new_password1')
        new_password2 = request.POST.get('new_password2')
        
        if new_password1 != new_password2:
            messages.error(request, '비밀번호가 일치하지 않습니다.')
            return render(request, 'crm_app/user_change_password.html', {'user': user})
        
        if len(new_password1) < 4:
            messages.error(request, '비밀번호는 최소 4자 이상이어야 합니다.')
            return render(request, 'crm_app/user_change_password.html', {'user': user})
        
        user.set_password(new_password1)
        user.save()
        
        messages.success(request, f'사용자 "{user.username}"의 비밀번호가 변경되었습니다.')
        return redirect('user_list')
    
    return render(request, 'crm_app/user_change_password.html', {'user': user})

